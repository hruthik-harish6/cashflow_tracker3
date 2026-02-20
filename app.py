from flask import Flask, render_template, request, redirect, session, url_for, flash, send_file, jsonify
import sqlite3
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
import os

app = Flask(__name__)
app.secret_key = "replace_this_with_a_random_secret_123"  # change this to a random string

DB_FILE = "tracker.db"

# --- Database init ---
def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS income (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            source TEXT,
            category TEXT,
            amount REAL NOT NULL,
            date TEXT,
            note TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
    ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS expense (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            item TEXT,
            category TEXT,
            amount REAL NOT NULL,
            date TEXT,
            note TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
    ''')

    conn.commit()
    conn.close()

init_db()

# --- Helpers ---
def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def login_required(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrap

# --- Auth ---
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        email = request.form.get("email","").strip().lower()
        password = request.form.get("password","")
        if not email or not password:
            flash("Email and password required.")
            return redirect(url_for("register"))
        hashed = generate_password_hash(password)
        conn = get_db(); c = conn.cursor()
        try:
            c.execute("INSERT INTO users (email, password) VALUES (?, ?)", (email, hashed))
            conn.commit()
            flash("Registered successfully. Log in.")
            return redirect(url_for("login"))
        except sqlite3.IntegrityError:
            flash("Email already registered.")
            return redirect(url_for("register"))
        finally:
            conn.close()
    return render_template("register.html")

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email","").strip().lower()
        password = request.form.get("password","")
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM users WHERE email = ?", (email,))
        user = c.fetchone()
        conn.close()
        if user and check_password_hash(user["password"], password):
            session["user_id"] = user["id"]
            session["email"] = user["email"]
            return redirect(url_for("home"))
        else:
            flash("Invalid credentials.")
            return redirect(url_for("login"))
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# --- Utility: build filter query ---
def build_filters(category, from_date, to_date):
    filters = []
    params = []

    if category and category != "All":
        # apply category to both income and expense when used later
        filters.append("category = ?")
        params.append(category)

    if from_date:
        filters.append("date >= ?")
        params.append(from_date)
    if to_date:
        filters.append("date <= ?")
        params.append(to_date)

    # return combined condition and params
    if filters:
        return " AND ".join(filters), params
    else:
        return "", []

# --- Home / Dashboard (supports GET filter query)
@app.route("/", methods=["GET"])
@login_required
def home():
    user_id = session["user_id"]

    # read filter params from query string
    category = request.args.get("category", default="All")
    from_date = request.args.get("from_date", default="")
    to_date = request.args.get("to_date", default="")

    conn = get_db(); c = conn.cursor()

    # Income query
    inc_query = "SELECT id, source, category, amount, date, note FROM income WHERE user_id = ?"
    inc_params = [user_id]
    if category and category != "All":
        inc_query += " AND category = ?"
        inc_params.append(category)
    if from_date:
        inc_query += " AND date >= ?"
        inc_params.append(from_date)
    if to_date:
        inc_query += " AND date <= ?"
        inc_params.append(to_date)
    inc_query += " ORDER BY date DESC, id DESC"
    c.execute(inc_query, inc_params)
    incomes = c.fetchall()

    # Expense query
    exp_query = "SELECT id, item, category, amount, date, note FROM expense WHERE user_id = ?"
    exp_params = [user_id]
    if category and category != "All":
        exp_query += " AND category = ?"
        exp_params.append(category)
    if from_date:
        exp_query += " AND date >= ?"
        exp_params.append(from_date)
    if to_date:
        exp_query += " AND date <= ?"
        exp_params.append(to_date)
    exp_query += " ORDER BY date DESC, id DESC"
    c.execute(exp_query, exp_params)
    expenses = c.fetchall()

    # totals (based on filtered results)
    total_income = sum([row["amount"] for row in incomes]) if incomes else 0.0
    total_expense = sum([row["amount"] for row in expenses]) if expenses else 0.0
    balance = total_income - total_expense

    conn.close()

    # Category list (income first then expense) — locked choices you asked
    categories = ["All", "Salary", "Business", "Freelance", "Others", "Food", "Travel", "Shopping", "Recharge", "Other"]

    return render_template("index.html",
                           incomes=incomes,
                           expenses=expenses,
                           balance=balance,
                           categories=categories,
                           selected_category=category,
                           from_date=from_date,
                           to_date=to_date,
                           email=session.get("email"))

# --- Add income / expense ---
@app.route("/add_income", methods=["POST"])
@login_required
def add_income():
    user_id = session["user_id"]
    source = request.form.get("source","").strip()
    category = request.form.get("category","").strip() or "Others"
    amount = float(request.form.get("amount","0") or 0)
    date_str = request.form.get("date") or datetime.now().strftime("%Y-%m-%d")
    note = request.form.get("note","").strip()

    conn = get_db(); c = conn.cursor()
    c.execute("INSERT INTO income (user_id, source, category, amount, date, note) VALUES (?,?,?,?,?,?)",
              (user_id, source, category, amount, date_str, note))
    conn.commit(); conn.close()
    return redirect(url_for("home"))

@app.route("/add_expense", methods=["POST"])
@login_required
def add_expense():
    user_id = session["user_id"]
    item = request.form.get("item","").strip()
    category = request.form.get("category","").strip() or "Other"
    amount = float(request.form.get("amount","0") or 0)
    date_str = request.form.get("date") or datetime.now().strftime("%Y-%m-%d")
    note = request.form.get("note","").strip()

    conn = get_db(); c = conn.cursor()
    c.execute("INSERT INTO expense (user_id, item, category, amount, date, note) VALUES (?,?,?,?,?,?)",
              (user_id, item, category, amount, date_str, note))
    conn.commit(); conn.close()
    return redirect(url_for("home"))

# --- Delete routes ---
@app.route("/delete_income/<int:id>")
@login_required
def delete_income(id):
    user_id = session["user_id"]
    conn = get_db(); c = conn.cursor()
    c.execute("DELETE FROM income WHERE id = ? AND user_id = ?", (id, user_id))
    conn.commit(); conn.close()
    return redirect(url_for("home"))

@app.route("/delete_expense/<int:id>")
@login_required
def delete_expense(id):
    user_id = session["user_id"]
    conn = get_db(); c = conn.cursor()
    c.execute("DELETE FROM expense WHERE id = ? AND user_id = ?", (id, user_id))
    conn.commit(); conn.close()
    return redirect(url_for("home"))

# --- Chart data for Chart.js ---
@app.route("/chart_data")
@login_required
def chart_data():
    user_id = session["user_id"]
    # optional filters from query params
    category = request.args.get("category", default="All")
    from_date = request.args.get("from_date", default="")
    to_date = request.args.get("to_date", default="")

    conn = get_db(); c = conn.cursor()

    # compute totals per category for expenses
    exp_query = "SELECT category, SUM(amount) as total FROM expense WHERE user_id = ?"
    params = [user_id]
    if category and category != "All":
        exp_query += " AND category = ?"
        params.append(category)
    if from_date:
        exp_query += " AND date >= ?"; params.append(from_date)
    if to_date:
        exp_query += " AND date <= ?"; params.append(to_date)
    exp_query += " GROUP BY category"
    c.execute(exp_query, params)
    exp_rows = c.fetchall()

    # income totals per category
    inc_query = "SELECT category, SUM(amount) as total FROM income WHERE user_id = ?"
    params2 = [user_id]
    if category and category != "All":
        inc_query += " AND category = ?"
        params2.append(category)
    if from_date:
        inc_query += " AND date >= ?"; params2.append(from_date)
    if to_date:
        inc_query += " AND date <= ?"; params2.append(to_date)
    inc_query += " GROUP BY category"
    c.execute(inc_query, params2)
    inc_rows = c.fetchall()

    conn.close()

    exp_labels = [r["category"] for r in exp_rows]
    exp_values = [r["total"] for r in exp_rows]
    inc_labels = [r["category"] for r in inc_rows]
    inc_values = [r["total"] for r in inc_rows]

    return jsonify({
        "expense": {"labels": exp_labels, "values": exp_values},
        "income": {"labels": inc_labels, "values": inc_values}
    })

# --- Export to Excel (.xlsx) ---
@app.route("/export")
@login_required
def export():
    user_id = session["user_id"]
    category = request.args.get("category", default="All")
    from_date = request.args.get("from_date", default="")
    to_date = request.args.get("to_date", default="")

    conn = get_db(); c = conn.cursor()

    # income
    inc_query = "SELECT id, source, category, amount, date, note FROM income WHERE user_id = ?"
    inc_params = [user_id]
    if category and category != "All":
        inc_query += " AND category = ?"; inc_params.append(category)
    if from_date:
        inc_query += " AND date >= ?"; inc_params.append(from_date)
    if to_date:
        inc_query += " AND date <= ?"; inc_params.append(to_date)
    inc_query += " ORDER BY date DESC"
    c.execute(inc_query, inc_params)
    inc_rows = c.fetchall()

    # expense
    exp_query = "SELECT id, item, category, amount, date, note FROM expense WHERE user_id = ?"
    exp_params = [user_id]
    if category and category != "All":
        exp_query += " AND category = ?"; exp_params.append(category)
    if from_date:
        exp_query += " AND date >= ?"; exp_params.append(from_date)
    if to_date:
        exp_query += " AND date <= ?"; exp_params.append(to_date)
    exp_query += " ORDER BY date DESC"
    c.execute(exp_query, exp_params)
    exp_rows = c.fetchall()

    conn.close()

    # create workbook
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Income"
    ws1.append(["ID", "Source", "Category", "Amount", "Date", "Note"])
    for r in inc_rows:
        ws1.append([r["id"], r["source"], r["category"], r["amount"], r["date"], r["note"]])

    ws2 = wb.create_sheet(title="Expense")
    ws2.append(["ID", "Item", "Category", "Amount", "Date", "Note"])
    for r in exp_rows:
        ws2.append([r["id"], r["item"], r["category"], r["amount"], r["date"], r["note"]])

    # save to bytes
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(bio, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
from flask import send_file
from openpyxl import Workbook
import io

@app.route('/export_excel')
def export_excel():
    conn = get_db_connection()
    incomes = conn.execute('SELECT * FROM income').fetchall()
    expenses = conn.execute('SELECT * FROM expenses').fetchall()
    conn.close()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Income"
    ws1.append(["ID", "Source", "Category", "Date", "Amount", "Note"])
    for row in incomes:
        ws1.append([row['id'], row['source'], row['category'], row['date'], row['amount'], row['note']])

    ws2 = wb.create_sheet("Expenses")
    ws2.append(["ID", "Item", "Category", "Date", "Amount", "Note"])
    for row in expenses:
        ws2.append([row['id'], row['item'], row['category'], row['date'], row['amount'], row['note']])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name="cashflow_data.xlsx")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0",port=port)

